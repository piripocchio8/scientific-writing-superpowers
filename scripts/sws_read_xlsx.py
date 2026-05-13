"""Read an .xlsx file as TSV — SWS wrapper around openpyxl.

Cycle #7 hotfix (D22): native Read tool does not parse .xlsx. SWS ships
this reader so drafting + budget-helper agents have a sanctioned path.
The --show-formulas flag preserves formula source, which the user's
xlsx-as-data-authority workflow (CLAUDE.md item #4) depends on.

CLI:
  sws_read_xlsx.py <file.xlsx>
  sws_read_xlsx.py <file.xlsx> --sheet "Data"
  sws_read_xlsx.py <file.xlsx> --sheet "Data" --range A1:C10
  sws_read_xlsx.py <file.xlsx> --sheet "Calc" --show-formulas

Exit codes:
  0  ok
  2  file not found (or invalid CLI arg, e.g. unknown sheet)
  3  openpyxl parse error
"""
from __future__ import annotations
import argparse
import sys
from pathlib import Path


def _format_cell(value) -> str:
    """Render a single cell to its TSV-safe string form."""
    if value is None:
        return ""
    return str(value)


def _emit_sheet(sheet, cell_range: str | None, with_header: bool, out=sys.stdout) -> None:
    """Print one worksheet (or a range within it) as TSV.

    with_header controls whether to print the sheet-name banner; default-mode
    multi-sheet prints set this True, single-sheet --sheet passes False.
    """
    if with_header:
        print(f"# Sheet: {sheet.title}", file=out)
    if cell_range:
        rows = sheet[cell_range]
        # sheet[range] returns a tuple-of-tuples for multi-cell ranges, or a
        # single Cell for a single coordinate. Normalize to a tuple-of-tuples.
        if not isinstance(rows, tuple):
            rows = ((rows,),)
        elif rows and not isinstance(rows[0], tuple):
            rows = (rows,)
        for row in rows:
            print("\t".join(_format_cell(c.value) for c in row), file=out)
    else:
        for row in sheet.iter_rows(values_only=True):
            print("\t".join(_format_cell(v) for v in row), file=out)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(
        description="Read .xlsx as TSV (SWS wrapper around openpyxl)."
    )
    ap.add_argument("file", help="Path to .xlsx file")
    ap.add_argument("--sheet", help="Limit to one sheet by name")
    ap.add_argument("--range", dest="cell_range",
                    help="Limit to a cell range, e.g. A1:B10")
    ap.add_argument("--show-formulas", action="store_true",
                    help="Print formula source for cells that contain formulas, "
                         "instead of cached values (data-authority workflow).")
    args = ap.parse_args(argv)

    path = Path(args.file)
    if not path.is_file():
        print(f"sws_read_xlsx: file not found: {path}", file=sys.stderr)
        return 2

    try:
        from openpyxl import load_workbook
        # data_only=False keeps formula source (=A1+B1) in cell.value; True
        # would replace it with the cached numeric value openpyxl read off
        # the file's calc cache.
        wb = load_workbook(str(path), data_only=not args.show_formulas)
    except Exception as exc:
        print(f"sws_read_xlsx: parse error: {exc}", file=sys.stderr)
        return 3

    if args.sheet:
        if args.sheet not in wb.sheetnames:
            print(f"sws_read_xlsx: sheet not found: {args.sheet!r}", file=sys.stderr)
            return 2
        _emit_sheet(wb[args.sheet], args.cell_range, with_header=False)
    else:
        if args.cell_range:
            print("sws_read_xlsx: --range requires --sheet", file=sys.stderr)
            return 2
        for name in wb.sheetnames:
            _emit_sheet(wb[name], None, with_header=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
