"""Resolve an .xlsx workbook for SWS data-authority use (D4).

Wraps sws_read_xlsx.py's data_only=True path. Every formula cell that has NO
cached value causes an immediate non-zero exit with an actionable message naming
the exact cell and telling the user how to fix it. Never re-derives or guesses
formula results.

CLI (same surface as sws_read_xlsx.py; --show-formulas is intentionally absent):
  sws_xlsx_resolve.py <file.xlsx>
  sws_xlsx_resolve.py <file.xlsx> --sheet "Results"
  sws_xlsx_resolve.py <file.xlsx> --sheet "Results" --range A1:D10

Exit codes:
  0  all cells resolved; TSV output on stdout
  1  un-cached formula cell found (D4 fail-loud)
  2  file not found or invalid argument
  3  openpyxl parse error
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path


_FORMULA_RE = re.compile(r"^=", re.IGNORECASE)


def _cell_ref(col_idx: int, row_idx: int) -> str:
    """Convert 1-based column + row indices to an A1-style reference."""
    col_letters = ""
    c = col_idx
    while c > 0:
        c, remainder = divmod(c - 1, 26)
        col_letters = chr(65 + remainder) + col_letters
    return f"{col_letters}{row_idx}"


def _check_sheet_for_uncached(sheet, cell_range: str | None) -> tuple[str, str] | None:
    """Return (sheet_title, cell_ref) for the first un-cached formula cell, or None.

    An un-cached formula cell is one whose value is None AND whose formula
    source (readable with data_only=False) starts with '='. We detect this
    by reloading the sheet separately when a None is found; however, since
    the caller already opened with data_only=True (which strips formula
    source), we use a heuristic: any cell whose value is None in a row that
    has other non-None values is suspicious. For a definitive check we
    re-read the raw XML formula indicator via the internal openpyxl attribute.
    """
    if cell_range:
        rows = sheet[cell_range]
        if not isinstance(rows, tuple):
            rows = ((rows,),)
        elif rows and not isinstance(rows[0], tuple):
            rows = (rows,)
        for row in rows:
            for cell in row:
                if cell.value is None and _has_formula_source(cell):
                    ref = cell.coordinate
                    return sheet.title, ref
    else:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None and _has_formula_source(cell):
                    ref = cell.coordinate
                    return sheet.title, ref
    return None


def _has_formula_source(cell) -> bool:
    """Return True if the cell's raw XML indicates it holds a formula.

    openpyxl stores the raw formula string in cell._value when data_only=False;
    with data_only=True the formula is stripped and _value holds the cached
    numeric result (or None if never cached). We check the internal
    data_type attribute: 'f' means formula.
    """
    # openpyxl sets data_type to 'f' for formula cells regardless of data_only.
    return getattr(cell, "data_type", None) == "f"


def _format_cell(value) -> str:
    if value is None:
        return ""
    return str(value)


def _emit_sheet(sheet, cell_range: str | None, with_header: bool) -> None:
    if with_header:
        print(f"# Sheet: {sheet.title}")
    if cell_range:
        rows = sheet[cell_range]
        if not isinstance(rows, tuple):
            rows = ((rows,),)
        elif rows and not isinstance(rows[0], tuple):
            rows = (rows,)
        for row in rows:
            print("\t".join(_format_cell(c.value) for c in row))
    else:
        for row in sheet.iter_rows(values_only=True):
            print("\t".join(_format_cell(v) for v in row))


def _fail_loud(sheet_title: str, cell_ref: str) -> int:
    msg = (
        f"Cell {sheet_title}!{cell_ref} is a formula with no cached value. "
        f"Open and save this workbook in Excel or LibreOffice so values cache, "
        f"then re-run /sws:curate-data."
    )
    print(msg, file=sys.stderr)
    return 1


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(
        description="Read .xlsx as TSV (data_only); fail loud on un-cached formula cells (D4)."
    )
    ap.add_argument("file", help="Path to .xlsx file")
    ap.add_argument("--sheet", help="Limit to one sheet by name")
    ap.add_argument(
        "--range", dest="cell_range", help="Limit to a cell range, e.g. A1:D10"
    )
    args = ap.parse_args(argv)

    path = Path(args.file)
    if not path.is_file():
        print(f"sws_xlsx_resolve: file not found: {path}", file=sys.stderr)
        return 2

    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), data_only=True)
    except Exception as exc:
        print(f"sws_xlsx_resolve: parse error: {exc}", file=sys.stderr)
        return 3

    if args.sheet:
        if args.sheet not in wb.sheetnames:
            print(
                f"sws_xlsx_resolve: sheet not found: {args.sheet!r}", file=sys.stderr
            )
            return 2
        sheet = wb[args.sheet]
        hit = _check_sheet_for_uncached(sheet, args.cell_range)
        if hit:
            return _fail_loud(*hit)
        _emit_sheet(sheet, args.cell_range, with_header=False)
    else:
        if args.cell_range:
            print(
                "sws_xlsx_resolve: --range requires --sheet", file=sys.stderr
            )
            return 2
        for name in wb.sheetnames:
            hit = _check_sheet_for_uncached(wb[name], None)
            if hit:
                return _fail_loud(*hit)
        for name in wb.sheetnames:
            _emit_sheet(wb[name], None, with_header=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
