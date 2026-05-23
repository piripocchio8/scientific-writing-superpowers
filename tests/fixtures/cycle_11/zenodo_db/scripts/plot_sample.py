"""Minimal plot script for smoke_cycle_11.sh fixture.

Reads Zenodo_db/data/sample.xlsx (sheet: Kinetics) and produces
Zenodo_db/figures/fig1_sample.png and fig1_sample.pdf.

This script is executed by sws_plot_runner.py (D6a). The runner injects
rcParams (font floor >= 8 pt, single-column width 7.5 cm) before execution.
The script itself does NOT set rcParams — it relies on the runner injection —
so it is compliant with the D6a font floor when called through the runner.
"""
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
ZENODO_DB = HERE.parent
DATA_XLSX = ZENODO_DB / "data" / "sample.xlsx"
OUT_PNG = ZENODO_DB / "figures" / "fig1_sample.png"
OUT_PDF = ZENODO_DB / "figures" / "fig1_sample.pdf"


def main():
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from openpyxl import load_workbook
    except ImportError as exc:
        print(f"plot_sample.py: missing dependency: {exc}", file=sys.stderr)
        sys.exit(2)

    wb = load_workbook(str(DATA_XLSX), data_only=True)
    ws = wb["Kinetics"]
    times = []
    concentrations = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            times.append(float(row[0]))
            concentrations.append(float(row[1]))

    # figsize uses the D6a single-column width (7.5 cm = 2.953 in); height free.
    fig, ax = plt.subplots(figsize=(2.953, 2.5))
    ax.plot(times, concentrations, marker="o", label="[compound]")
    ax.set_xlabel("Time (s)")
    ax.set_ylabel("Concentration (mM)")
    ax.set_title("Sample kinetics")
    ax.legend()

    fig.savefig(str(OUT_PNG), dpi=300, bbox_inches="tight")
    fig.savefig(str(OUT_PDF), bbox_inches="tight")
    plt.close(fig)
    print(f"plot_sample.py: wrote {OUT_PNG} and {OUT_PDF}")


if __name__ == "__main__":
    main()
