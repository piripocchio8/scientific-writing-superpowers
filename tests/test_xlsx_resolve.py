"""Unit tests for sws_xlsx_resolve.py — D4 fail-loud on un-cached formula cells.

All fixtures are built inline with openpyxl; no binary xlsx files committed.
"""
from __future__ import annotations

import importlib.util
import io
import subprocess
import sys
from contextlib import redirect_stderr
from pathlib import Path

import pytest

try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

REPO = Path(__file__).resolve().parents[1]
SCRIPT = REPO / "scripts" / "sws_xlsx_resolve.py"

# openpyxl-dependent fixture tests carry this skip marker individually; the
# module is intentionally NOT marked at module level so the platform-independent
# D4 unit test below always runs.
_needs_openpyxl = pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")


def _load_resolver_module():
    """Import sws_xlsx_resolve as a module so its internals can be unit-tested.

    Loaded by file path to avoid relying on openpyxl (the module only imports
    openpyxl lazily inside main()), keeping this platform-independent.
    """
    spec = importlib.util.spec_from_file_location("sws_xlsx_resolve_under_test", SCRIPT)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


class _StubCell:
    """Minimal stand-in for an openpyxl cell: a None value + data_type='f'."""

    def __init__(self, value, data_type, coordinate):
        self.value = value
        self.data_type = data_type
        self.coordinate = coordinate


class _StubSheet:
    """Minimal stand-in for an openpyxl worksheet exposing iter_rows()."""

    def __init__(self, title, rows):
        self.title = title
        self._rows = rows

    def iter_rows(self):
        return iter(self._rows)


# ---------------------------------------------------------------------------
# Platform-independent fail-loud unit test (no openpyxl round-trip required).
# This MUST run on every platform — it exercises the D4 detection + fail-loud
# branch by driving the resolver internals directly with stub cells, which is
# what openpyxl's round-trip cannot reliably reproduce here.
# ---------------------------------------------------------------------------

def test_d4_detection_and_fail_loud_verbatim_message():
    mod = _load_resolver_module()

    # (a) detection: a cell with data_type='f' and value=None is flagged.
    uncached = _StubCell(value=None, data_type="f", coordinate="C1")
    assert mod._has_formula_source(uncached) is True

    # A formula cell with a CACHED value (value not None) must NOT be flagged.
    cached = _StubCell(value=42.0, data_type="f", coordinate="C2")
    cached_only_sheet = _StubSheet("Calc", [[cached]])
    assert mod._check_sheet_for_uncached(cached_only_sheet, None) is None

    # The un-cached formula cell IS flagged, naming sheet + cell.
    sheet = _StubSheet("Results", [[cached, uncached]])
    hit = mod._check_sheet_for_uncached(sheet, None)
    assert hit == ("Results", "C1"), f"expected detection of Results!C1, got {hit}"

    # (b) fail-loud: non-zero exit code + VERBATIM message on stderr.
    buf = io.StringIO()
    with redirect_stderr(buf):
        rc = mod._fail_loud(*hit)
    assert rc == 1, f"fail-loud must return exit code 1, got {rc}"
    expected = (
        "Cell Results!C1 is a formula with no cached value. "
        "Open and save this workbook in Excel or LibreOffice so values cache, "
        "then re-run /sws:curate-data."
    )
    assert buf.getvalue().strip() == expected, (
        f"verbatim D4 message mismatch:\n got: {buf.getvalue().strip()!r}\n exp: {expected!r}"
    )


# ---------------------------------------------------------------------------
# Fixture builders
# ---------------------------------------------------------------------------

def _make_clean_workbook(tmp_path: Path) -> Path:
    """Workbook with only plain numeric values — no formulas anywhere."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "compound"
    ws["B1"] = "yield_pct"
    ws["A2"] = "cpd-1"
    ws["B2"] = 72.5
    ws["A3"] = "cpd-2"
    ws["B3"] = 65.0
    path = tmp_path / "clean.xlsx"
    wb.save(str(path))
    return path


def _make_workbook_with_cached_formula(tmp_path: Path) -> Path:
    """Workbook where a formula cell HAS a cached numeric value.

    openpyxl data_only=True will return the cached number, not None.
    We simulate this by saving the workbook with data_only=False (formula
    source preserved), then reloading it with data_only=True to confirm the
    cached value is accessible. Since openpyxl cannot truly cache values at
    write-time the same way Excel does, we use a plain value cell to represent
    a "cached" formula result — the resolver must accept it.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws["A1"] = 10.0
    ws["B1"] = 32.0
    # Represent cached formula result as a plain numeric value.
    # (openpyxl cannot inject Excel calc-cache entries; the test covers the
    # code path where data_only read returns a non-None numeric value.)
    ws["C1"] = 42.0
    path = tmp_path / "cached.xlsx"
    wb.save(str(path))
    return path


def _make_workbook_with_uncached_formula(tmp_path: Path) -> Path:
    """Workbook where a formula cell has NO cached value.

    openpyxl data_only=True on a freshly written workbook returns None for
    formula cells because Excel never ran the calculation. This is the
    fail-loud path.
    """
    from openpyxl import load_workbook as lw
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws["A1"] = 5.0
    ws["B1"] = 3.0
    ws["C1"] = "=A1+B1"  # formula source; no cached value
    path = tmp_path / "uncached.xlsx"
    wb.save(str(path))
    # Confirm openpyxl data_only=True truly returns None for C1.
    wb2 = lw(str(path), data_only=True)
    cell = wb2["Results"]["C1"]
    cell_val = cell.value
    if cell_val is not None:
        pytest.skip("openpyxl cached the formula value on this platform — skip fail-loud test")
    if getattr(cell, "data_type", None) != "f":
        pytest.skip(
            "openpyxl lost data_type='f' on data_only load on this platform — "
            "fail-loud detection requires app-saved workbooks; skip fixture-based test"
        )
    return path


def _run(args, expect_zero: bool = True):
    cp = subprocess.run(
        [sys.executable, str(SCRIPT), *args],
        capture_output=True, text=True,
    )
    if expect_zero:
        assert cp.returncode == 0, f"unexpected non-zero: {cp.stderr}"
    return cp


# ---------------------------------------------------------------------------
# Happy paths
# ---------------------------------------------------------------------------

@_needs_openpyxl
def test_clean_workbook_exits_zero_and_emits_tsv(tmp_path):
    path = _make_clean_workbook(tmp_path)
    cp = _run([str(path)])
    assert "compound" in cp.stdout
    assert "cpd-1" in cp.stdout
    assert "72.5" in cp.stdout


@_needs_openpyxl
def test_cached_value_workbook_exits_zero(tmp_path):
    path = _make_workbook_with_cached_formula(tmp_path)
    cp = _run([str(path)])
    assert cp.returncode == 0


@_needs_openpyxl
def test_sheet_scope_respected(tmp_path):
    path = _make_clean_workbook(tmp_path)
    cp = _run([str(path), "--sheet", "Data"])
    assert "compound" in cp.stdout
    assert cp.returncode == 0


@_needs_openpyxl
def test_range_scope_respected(tmp_path):
    path = _make_clean_workbook(tmp_path)
    cp = _run([str(path), "--sheet", "Data", "--range", "A2:B3"])
    assert "cpd-1" in cp.stdout
    assert "compound" not in cp.stdout  # header row excluded
    assert cp.returncode == 0


# ---------------------------------------------------------------------------
# Fail-loud path (D4)
# ---------------------------------------------------------------------------

@_needs_openpyxl
def test_uncached_formula_cell_exits_nonzero(tmp_path):
    path = _make_workbook_with_uncached_formula(tmp_path)
    cp = _run([str(path)], expect_zero=False)
    assert cp.returncode != 0


@_needs_openpyxl
def test_uncached_formula_message_names_sheet_and_cell(tmp_path):
    path = _make_workbook_with_uncached_formula(tmp_path)
    cp = _run([str(path)], expect_zero=False)
    # D4: message must name the sheet and cell reference
    assert "Results" in cp.stderr or "Results" in cp.stdout
    assert "C1" in cp.stderr or "C1" in cp.stdout


@_needs_openpyxl
def test_uncached_formula_message_instructs_open_save(tmp_path):
    """D4: actionable message must mention open + save and /sws:curate-data."""
    path = _make_workbook_with_uncached_formula(tmp_path)
    cp = _run([str(path)], expect_zero=False)
    combined = cp.stderr + cp.stdout
    assert "Open and save" in combined or "open and save" in combined
    assert "/sws:curate-data" in combined


@_needs_openpyxl
def test_uncached_formula_verbatim_message_fragment(tmp_path):
    """D4: the exact message fragment from the spec must appear verbatim."""
    path = _make_workbook_with_uncached_formula(tmp_path)
    cp = _run([str(path)], expect_zero=False)
    combined = cp.stderr + cp.stdout
    assert "is a formula with no cached value" in combined


@_needs_openpyxl
def test_missing_file_exits_2(tmp_path):
    cp = _run([str(tmp_path / "nonexistent.xlsx")], expect_zero=False)
    assert cp.returncode == 2


@_needs_openpyxl
def test_malformed_file_exits_3(tmp_path):
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"not a real xlsx file")
    cp = _run([str(bad)], expect_zero=False)
    assert cp.returncode == 3
